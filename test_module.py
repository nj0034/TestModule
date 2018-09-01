import uuid
from sailer.sailer import Sailer
from sailer.utils import *
import random
import time
import re
import urllib.parse
import json
from elasticsearch import Elasticsearch
import boto3
import os
import requests
from bs4 import BeautifulSoup
from PIL import Image
from openpyxl import load_workbook

json_data = open('json/메인.json', encoding='UTF8').read()
XPATH_JSON = json.loads(json_data)
es = Elasticsearch('https://search-toast-test-4gvyyphmm2klzciadaeqkkzkza.ap-northeast-2.es.amazonaws.com')
S3_ENDPOINT = "https://s3.ap-northeast-2.amazonaws.com/toast-luna-dev/{bucket}/{type}/{filename}"
END_PAGE = 100
GRAPHQL_ENDPOINT = "https://luna.toast.one/api/v2/"


class TestModule(Sailer):
    def start(self):
        self.row = 2
        # result = es.get(index='parsing-json-skku', doc_type='_doc', id=id)
        #
        # if result['found']:
        #     XPATH_JSON = result['_source']['json']
        # else:
        #     XPATH_JSON = None

        self.meta = XPATH_JSON['meta']
        self.parser = XPATH_JSON['parser']
        self.rule = XPATH_JSON['rule']
        self.props = self.rule['properties']
        self.bucket = self.meta['type']

        method = self.parser['method']
        page_url_list = self.rule.get('page_url_list')

        if not page_url_list:
            page_url_list = [self.rule['page_url']]

        for page_url in page_url_list:
            if method == 'url_based':
                start_page = int(self.rule['start_page'])
                page_increase = int(self.rule['page_increase'])

                for i in range(start_page, start_page + END_PAGE):
                    self.go(page_url.format(page=i * page_increase))
                    print("# {} page start".format(i))
                    self.url_based()

            elif method == 'next_button':
                self.go(page_url)
                self.next_button()

            elif method == 'scroll_down':
                self.go(page_url)
                self.xpath(r'//*[@id="bar-sorter"]/div/button/span').click()
                time.sleep(1)
                self.xpath(r'//*[@id="bar-sorter"]/div/ul/li[4]/label').click()

                for i in range(20):
                    print("down")
                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(2)
                self.scroll_down()

            else:
                pass

    def scroll_down(self):
        # parsing_result_json_list = self.parsing_title_url()
        # print(parsing_result_json_list)
        self.sheet = "contest"

        wb = load_workbook('링커리어.xlsx')
        ws = wb.create_sheet(title=self.sheet)

        elements = self.driver.find_elements_by_class_name('wrapper_content_text')

        for i, element in enumerate(elements):
            result = re.compile(r'(.*)\s조회\s(\d+)회\s+댓글 (\d+)개').search(element.text)
            print(result.group(1), result.group(2), result.group(3))
            ws.cell(row=i + 2, column=1, value=result.group(1))
            ws.cell(row=i + 2, column=2, value=result.group(2))
            ws.cell(row=i + 2, column=3, value=result.group(3))

        wb.save('링커리어.xlsx')
        wb.close()
        self.sheet = "activity"

    def next_button(self):
        in_prop_list = self.props.keys()
        self.wait_xpath(self.rule['top_article_xpath'])
        self.xpath(self.rule['top_article_xpath']).click()

        while (True):
            parsing_result_json_list = [{"url": self.current_url}]
            self.parsing_in_props(in_prop_list, parsing_result_json_list)
            # print(parsing_result_json_list)
            #
            # # es에 parsing_result_json_list 저장(top 글이면 저장 안함)
            # self.store_in_es(parsing_result_json_list[0])

            next_button_url = self.xpath(self.rule['next_button_xpath']).get_attribute('href')
            if next_button_url:
                self.go(next_button_url)
            else:
                break

    def url_based(self):
        out_prop_list = list()
        in_prop_list = list()
        for key, value in self.props.items():
            if value.get('position') == 'out':
                out_prop_list.append(key)
            elif value.get('position') == 'in':
                in_prop_list.append(key)
            else:
                pass

        parsing_result_json_list = self.parsing_title_url()

        # 리스트에서 파싱해야 하는 props
        for out_prop in out_prop_list:
            web_elements = self.xpaths(self.props[out_prop]['xpath'])
            for web_element, parsing_result_json in zip(web_elements, parsing_result_json_list):
                out_prop_json = self.parsing_prop(prop=out_prop, web_element=web_element)
                parsing_result_json.update(out_prop_json)

        # 내부 prop 파싱 후 es에 저장
        self.parsing_in_props(in_prop_list, parsing_result_json_list)

    def parsing_title_url(self):
        title_url_json = self.props['title_url']

        if title_url_json.get('xpath'):
            title_url_web_elements = self.xpaths(title_url_json['xpath'])
            href_list = [title_url_web_element.get_attribute('href') for title_url_web_element in
                         title_url_web_elements]

            # 자바스크립트 url
            # href 안에서 파라미터만 추출해서 base_url form 에 넣어서 생성
            param_regex = title_url_json.get('param_regex')
            if param_regex:
                base_url = self.props['title_url']['base_url']
                title_url_list = self.combine_url_with_params(param_regex, href_list, base_url)

                parsing_result_json_list = [{"url": title_url} for title_url in title_url_list]

            # 보통 url
            else:
                parsing_result_json_list = [{"url": href} for href in href_list]

        # regex로 url 파싱
        else:
            regex = self.props['title_url']['regex']
            html = self.driver.find_element_by_class_name(title_url_json['html_class_name']).get_attribute('innerHTML')
            title_url_list = re.findall(regex, html)
            parsing_result_json_list = [{"url": title_url} for title_url in title_url_list]

        return parsing_result_json_list

    def parsing_in_props(self, in_prop_list, parsing_result_json_list):
        for parsing_result_json in parsing_result_json_list:
            self.go(parsing_result_json['url'])
            for in_prop in in_prop_list:
                regex = self.props[in_prop].get('regex')
                class_name = self.props[in_prop].get('class_name')
                xpath = self.props[in_prop].get('xpath')
                # 캠퍼즈
                html_class_name = self.props[in_prop].get('html_class_name')

                # 정규표현식
                if regex:
                    if xpath:
                        html = self.xpath(xpath).get_attribute('innerHTML')
                    else:
                        html = self.driver.find_element_by_class_name(html_class_name).get_attribute('innerHTML')
                    in_prop_json = self.parsing_regex_prop(in_prop, regex, html)

                # class name
                elif class_name:
                    web_element = self.driver.find_element_by_class_name(class_name)
                    in_prop_json = self.parsing_prop(prop=in_prop, web_element=web_element)

                # xpath
                else:
                    in_prop_json = self.parsing_prop(prop=in_prop)

                parsing_result_json.update(in_prop_json)

            # store in es
            self.store_in_es(parsing_result_json)

    def parsing_prop(self, **kwargs):
        prop = kwargs.get('prop', '')
        prop_info = self.props[prop]

        xpath = prop_info.get('xpath')

        web_element = kwargs.get('web_element')
        web_elements = None

        if not web_element:
            try:
                self.wait_xpath(xpath)
                web_element = self.xpath(xpath)
                web_elements = self.xpaths(xpath)
            except:
                web_element = None
                web_elements = None

        prop_type = prop_info['type']

        if web_element or web_elements:
            if prop_type == 'content':
                prop_json = {
                    "content_text": web_element.text,
                    "content_HTML": web_element.get_attribute('innerHTML'),
                }

            elif prop_type == 'date':
                date = web_element.text
                date = date.replace('오전', 'AM')
                date = date.replace('오후', 'PM')

                attach = prop_info.get('attach')
                if attach:
                    date = attach + date

                format = prop_info['format']
                date = convert_datetime(date, format, '%Y-%m-%d %H:%M:%S')
                prop_json = {"created_datetime": date}

            elif prop_type == 'image':
                img_url_list = [web_element.get_attribute('src') for web_element in web_elements if web_element]

                prop_json = {"image_url_list": img_url_list}

            elif prop_type == 'poster':
                if prop_info['url'] == "href":
                    attr = "href"
                else:
                    attr = "src"

                poster_url = web_element.get_attribute(attr)
                poster = self.download_to_s3(bucket=self.bucket, type="poster", filename=uuid.uuid4(), url=poster_url)

                prop_json = {"poster": poster, "thumbnail": "_".join([poster, "thumbnail"])}

            elif prop_type == 'file':
                attach_name_list = list()
                href_list = [urllib.parse.unquote(web_element.get_attribute('href')) for web_element in web_elements if
                             web_element]

                param_regex = prop_info.get('param_regex')
                if param_regex:
                    base_url = prop_info['base_url']
                    attach_url_list = self.combine_url_with_params(param_regex, href_list, base_url)

                else:
                    attach_url_list = href_list

                if attach_url_list:
                    if prop_info.get('splited', ''):
                        attach_name_list = [
                            re.compile(prop_info['name_regex']).search(web_element.get_attribute('innerHTML')).group(1)
                            for web_element in self.xpaths(prop_info['HTML_xpath'])]
                    else:
                        attach_name_list = re.findall(prop_info['name_regex'],
                                                      self.xpath(prop_info['HTML_xpath']).get_attribute('innerHTML'))

                prop_json = {
                    prop: [{"url": self.download_to_s3(bucket=self.bucket, type="attach", filename=name, url=url),
                            "name": name} for
                           url, name in zip(attach_url_list, attach_name_list)]
                }

            elif prop_type == "url":
                prop_json = {prop: web_element.get_attribute('href')}

            else:
                prop_json = {prop: web_element.text}
            print(prop_json)

        else:
            prop_json = {}
            print(prop, "has no element")

        return prop_json

    def parsing_regex_prop(self, prop, regex, html):
        prop_data = re.compile(regex).search(html).group(1)

        if self.props[prop]['type'] == 'date':
            format = self.props[prop]['format']
            prop_data = convert_datetime(prop_data, format, '%Y-%m-%d %H:%M:%S')

        prop_json = {prop: prop_data}

        if self.props[prop]['type'] == 'poster':
            poster = self.download_to_s3(bucket=self.bucket, type="poster", filename=uuid.uuid4(), url=prop_data)
            prop_json = {"poster": poster, "thumbnail": "_".join([poster, "thumbnail"])}

        print(prop_json)
        return prop_json

    def convert_notice_img_url(self, result_json):
        image_url_list = result_json.get('image_url_list', '')
        if image_url_list:
            filename = str(uuid.uuid4())
            img_s3_url_list = [
                self.download_to_s3(bucket=self.bucket, type="image", filename="_".join([filename, str(i)]),
                                    url=img_url) for i, img_url in enumerate(image_url_list)]

            str_html = result_json['content_HTML']
            str_html = str_html.replace('&amp;', '&')
            html = BeautifulSoup(str_html, "html.parser")
            img_src_list = [img.get('src') for img in html.find_all('img')]

            for img_src, img_s3_url in zip(img_src_list, img_s3_url_list):
                str_html = str_html.replace(img_src, img_s3_url)

            result_json['content_HTML'] = str_html
            del result_json['image_url_list']

        return result_json

    def download_to_s3(self, **kwargs):
        filename = kwargs.get('filename', '')
        url = kwargs.get('url', '')
        try:
            if not filename:
                filename = url.split('/')[-1]

            filepath = r'./tmp/%s' % filename
            res = requests.get(url, stream=True)
            with open(filepath, "wb") as file:
                for chunk in res:
                    file.write(chunk)

            s3 = boto3.resource('s3')

            if kwargs['type'] == "poster":
                thumbnail_filepath = self.resize_thumbnail(filepath)
                s3.meta.client.upload_file(thumbnail_filepath, 'toast-luna-dev',
                                           '{bucket}/{type}/{filename}_thumbnail'.format(**kwargs))
                os.remove(thumbnail_filepath)

            s3.meta.client.upload_file(filepath, 'toast-luna-dev', '{bucket}/{type}/{filename}'.format(**kwargs))
            s3_url = S3_ENDPOINT.format(**kwargs)

            os.remove(filepath)
            return s3_url
        except:
            return None

    def store_in_db(self, result_json):
        result_json.update({"source": self.meta['site_name']})
        headers = {"Content-Type": "application/graphql"}
        data = """
        mutation {{
        createNotice(
        title: "{title}",
        createdDatetime: "{created_datetime}",
        contentText: "{content_text}",
            contentHtml: "{content_HTML}",
            url: "{url}",
            source: "{source}"
            attaches: {attach},
            id
            title
          }}
        }}""".format(**result_json).encode('utf8')
        print(data)
        result = requests.post(GRAPHQL_ENDPOINT, data=data, headers=headers)
        print(result.text)

    # es에 저장

    def store_in_es(self, result_json):
        self.store_in_db(result_json)
        # if self.store_in_excel(result_json):
        #     return

        check_result = self.check_duplication(result_json)
        if not check_result:
            result_json = self.convert_notice_img_url(result_json)
            result_json.update({"source": self.meta['site_name'], "hit": 0})
            try:
                es.create(index=self.meta['ES_index'], doc_type="_doc", id=uuid.uuid4(), body=result_json)
            except Exception as e:
                error_body = {"url": result_json['url'], "title": result_json['title'], "error_message": str(e)}
                es.create(index="token_error", doc_type="_doc", id=uuid.uuid4(), body=error_body)

        else:
            print("duplicate!")

        time_interval = [int(n) for n in self.parser['interval'].split('-')]
        random_time = random.randint(*time_interval)
        time.sleep(random_time)

    def check_duplication(self, result_json):
        title = result_json['title']
        source = self.meta['site_name']
        query_body = {"query": {"bool": {"must": [{"term": {"title.keyword": title}}, {"term": {"source": source}}]}}}
        result = es.search(index=self.meta['ES_index'], body=query_body)

        return result['hits']['hits']

    @staticmethod
    def combine_url_with_params(param_regex, href_list, base_url):
        url_list = list()
        param_list = [re.findall(param_regex, href) for href in href_list]

        for param in param_list:
            param = param[0]
            if isinstance(param, tuple):
                url_list.append(base_url.format(*param))
            else:
                url_list.append(base_url.format(param))

        return url_list

    @staticmethod
    def resize_thumbnail(filepath):
        thumbnail_filepath = "_".join([filepath, "thumbnail.jpg"])
        thumbnail_img = Image.open(filepath)
        new_width = 400
        wpercent = (new_width / float(thumbnail_img.size[0]))
        new_height = int((float(thumbnail_img.size[1]) * float(wpercent)))
        thumbnail_img.thumbnail((new_width, new_height), Image.ANTIALIAS)
        thumbnail_img = thumbnail_img.convert("RGB")
        thumbnail_img.save(thumbnail_filepath)

        return thumbnail_filepath

    def store_in_excel(self, result_json):
        print(self.row)
        wb = load_workbook('notice.xlsx')
        # ws = wb.create_sheet(title="main notice")
        ws = wb.active

        ws.cell(row=self.row, column=1, value=result_json['title'])
        ws.cell(row=self.row, column=2, value=result_json['content_text'])
        ws.cell(row=self.row, column=3, value=result_json['url'])

        wb.save('notice.xlsx')
        wb.close()
        self.row += 1

        return self.row


test_module = TestModule()
test_module.start()
test_module.close()
